
# pylint: disable=line-too-long, broad-exception-caught
'''
Script load_hospital_costs.py

A python script to the hospital costs data for a hospital
into the Clinical Costing tables, from an Excel workbook.

NOTE: The hospital activity must be loaded before the hospital costs,
as itemized costs refer to episodes of activity.

    SYNOPSIS:
    $ python load_hospital_costs.py
        [-D DatabaseType|--DatabaseType=DatabaseType]
        [-C configDir|--configDir=configDir]
        [-c configFile|--configFile=configFile]
        [-I inputDir|--inputDir=inputDir]
        [-i inputWorkbook|--inputWorkbook=inputWorkbook]
        [-s server|--server=server]
        [-u username|--username=username]
        [-p password|--password=password]
        [-d databaseName|--databaseName=databaseName]
        [-v loggingLevel|--verbose=logingLevel]
        [-L logDir|--logDir=logDir]
        [-l logfile|--logfile=logfile]

    REQUIRED
    -D DatabaseType|--DatabaseType=DatabaseType
    The type of database [choice:MSSQL/MySQL]


    OPTIONS
    -C configDir|--configDir=configDir
    The directory containing the database connection configuration file
    (default='databaseConfig')

    -c configFile|--configFile=configFile
    The database connection configuration file (default=clinical_costing.json)
    which has the default database values for each Database Type.
    These can be overwritten using command line options.

    -I inputDir|--inputDir=inputDir
    The directory containing the Excel workbook which contains
    the hospital patient activity data to be loaded.

    -i inputWorkbook|--inputWorkbook=inputWorkbook
    The Excel workbook containing the hospital patient activity data to be loaded.

    -s server|--server=server]
    The address of the database server

    -u userName|--userName=userName]
    The user name require to access the database

    -p password|--userName=userName]
    The user password require to access the database

    -d databaseName|--databaseName=databaseName]
    The name of the database

    -v loggingLevel|--verbose=loggingLevel
    Set the level of logging that you want.

    -O logDir|--logDir=logDir
    The directory where the log file will be created (default=".").

    -o logfile|--logfile=logfile
    The name of a log file where you want all messages captured.


    THE MAIN CODE
    Start by parsing the command line arguements, setting up logging
    and connecting to the database.
    Then check the Excel workbook - it should have one sheet for the hospital code/name,
    one sheet for the model code/name and one sheet listing the feeder data worksheets
    and their matching feeder_code, followed by the feeder data worksheets.
    And feeder data each sheet should have a header row with headings
    that match the names of the columns in the itemized_data database table.
    There should also be one worksheet of general ledger run adjustments and
    one worksheet of gl attributes run adjustments.
    And each sheet should have a header row with headings
    that match the names of the columns in the associated database table.
'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, unspecified-encoding

import os
import sys
import argparse
import logging
import datetime
import pandas as pd
from sqlalchemy import text, select, delete
from openpyxl import load_workbook
import functions as f
import data as d


if __name__ == '__main__':
    '''
    The main code
    Start by parsing the command line arguements, setting up logging
    and connecting to the database.
    Then check the Excel workbook - it should have one sheet for the hospital code/name,
    one sheet for the model code/name and one sheet listing the feeder data worksheets
    and their matching feeder_code, followed by the feeder data worksheets.
    And feeder data each sheet should have a header row with headings
    that match the names of the columns in the itemized_data database table.
    There should also be one worksheet of general ledger run adjustments and
    one worksheet of gl attributes run adjustments.
    And each sheet should have a header row with headings
    that match the names of the columns in the associated database table.
    '''

    # Save the program name
    progName = sys.argv[0]
    progName = progName[0:-3]        # Strip off the .py ending

    # Set the options
    parser = argparse.ArgumentParser(description='Load a Clinical Costing Model')
    parser.add_argument('-I', '--inputDir', dest='inputDir', default='.',
                        help='The directory containing the Excel workbook which contains the hospital patient activity data to be loaded.')
    parser.add_argument('-i', '--inputWorkbook', dest='inputWorkbook',
                        default='.', help='The name of the Excel workbook containing the hospital patient activity data to be loaded')
    f.addCommonArguments(parser)      # Add the common command line arguments
    args = parser.parse_args()

    # Parse the command line options
    args = parser.parse_args()
    inputDir = args.inputDir
    inputWorkbook = args.inputWorkbook
    configDir = args.configDir
    configFile = args.configFile
    DatabaseType = args.DatabaseType
    server = args.server
    username = args.username
    password = args.password
    databaseName = args.databaseName
    logDir = args.logDir
    logFile = args.logFile
    loggingLevel = args.verbose

    # Set up logging
    f.setupLogging(progName, logDir, logFile, loggingLevel)

    # Read in the configuration file - which must exist if required - and create the database engine
    f.createEngine(configDir, configFile, DatabaseType, server, username, password, databaseName)

    # Load the workbook
    wb = load_workbook(os.path.join(inputDir, inputWorkbook))

    # Check the 'hospital' worksheet
    table_df = f.checkWorksheet(wb, 'hospital', 'hospitals', [])

    # Get the hospital_code from the 'hospital' worksheet
    ws = wb['hospital']
    d.hospital_code = ws['A2'].value        # First (and only) hospital_code in the list

    # Check if this is a new hospital code, or upgraded configuration of an existing hospital
    hospitals_df = pd.read_sql_query(text('SELECT hospital_code FROM hospitals'), d.engine.connect())
    hospitals = hospitals_df.values.tolist()      # convert rows/columns to a list of lists (will be [[hospital_code]] )
    if not [d.hospital_code] in hospitals:
        logging.critical('hospital (%s) no in table "hospitals"', d.hospital_code)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)

    # Check the 'run' worksheet
    run_df = f.checkWorksheet(wb, 'run', 'clinical_costing_runs', ['hospital_code'])

    # Get the run_code from the 'run' worksheet
    ws = wb['run']
    d.run_code = ws['A2'].value        # First (and only) run_code in the list

    # Create the general "where" clause
    where = 'hospital_code = "' + d.hospital_code + '" AND run_code = "' + d.run_code + '"'

    # Check that we have an 'itemized costs' worksheet
    if 'itemized costs' not in wb:
        logging.critical('Missing "itemized costs" worksheet')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)

    # Check each itemized costs worksheet
    ws = wb['itemized costs']
    data = ws.values
    cols = next(data)
    itemized_costs_df = pd.DataFrame(list(data), columns=cols)
    if 'worksheet' not in itemized_costs_df.columns:
        logging.critical('Missing "worksheet" heading in "itemized costs" worksheet')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if 'feeder_code' not in itemized_costs_df.columns:
        logging.critical('Missing "feeder_code" heading in "itemized costs" worksheet')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    sheet_table_df = {}
    for sheet in itemized_costs_df['worksheet']:
        sheet_table_df[sheet] = f.checkWorksheet(wb, sheet, 'itemized_costs', ['hospital_code', 'run_code', 'feeder_code'])
        # Now check that the episode numbers are valid
        for row in sheet_table_df[sheet].itertuples(index=False):
            service_code = getattr(row, 'service_code')
            episode_no = getattr(row, 'episode_no')
            thisWhere = where + ' AND episode_no = "' + str(episode_no) + '"'
            found = True
            if service_code == 'Inpat':
                with d.Session() as session:
                    results = session.scalars(select(d.metadata.tables['inpat_episode_details']).where(text(thisWhere))).all()
                    if len(results) == 0:
                        found = False
            elif service_code == 'Clinic':
                with d.Session() as session:
                    results = session.scalars(select(d.metadata.tables['clinic_activity_details']).where(text(thisWhere))).all()
                    if len(results) == 0:
                        found = False
            elif service_code == 'ED':
                with d.Session() as session:
                    results = session.scalars(select(d.metadata.tables['ed_episode_details']).where(text(thisWhere))).all()
                    if len(results) == 0:
                        found = False
            else:
                logging.critical('Invalid service_code (%s) in "itemized costs" worksheet(%s)', service_code, sheet)
                logging.shutdown()
                sys.exit(d.EX_CONFIG)
            if not found:
                logging.critical('Invalid episode_no (%s) for service (%s) in item costs worksheet (%s)', episode_no, service_code, sheet)
                logging.shutdown()
                sys.exit(d.EX_CONFIG)

    # Check the 'general ledger run adjustments' worksheet
    costAdjustments_df = f.checkWorksheet(wb, 'general ledger run adjustments', 'general_ledger_run_adjustments', ['hospital_code', 'run_code'])

    # Check the 'general ledger run adjustments' worksheet
    attributeAdjustments_df = f.checkWorksheet(wb, 'gl attributes run adjustments', 'gl_attributes_run_adjustments', ['hospital_code', 'run_code'])

    # Check if this is a run_code exists
    runs_df = pd.read_sql_query(text('SELECT run_code FROM clinical_costing_runs'), d.engine.connect())
    runs = runs_df.values.tolist()      # convert rows/columns to a list of lists (will be [[run_code]] )
    newRun = not [d.run_code] in runs

    # Check the 'general ledger costs' worksheet
    general_ledger_table_df = f.checkWorksheet(wb, 'general ledger costs', 'general_ledger_costs', ['hospital_code', 'run_code'])

    # If this is a new run_code then add it to the table
    run_df = run_df.truncate(after=0)       # We only want the first row
    if newRun:
        # Prepend the hospital code and create a new run record
        run_df.insert(0,'hospital_code', d.hospital_code)
        run_df.to_sql('clinical_costing_runs', d.engine, if_exists='append', index=False)
    else:       # Check that this is the same run
        runs_df = pd.read_sql_query(text('SELECT * FROM clinical_costing_runs'), d.engine.connect())
        thisRun_df = runs_df[runs_df['run_code'] == d.run_code]
        ws = wb['run']
        run_description = ws['B2'].value
        start_date = ws['C2'].value
        if isinstance(start_date, datetime.datetime):
            start_date = start_date.date()
        end_date = ws['D2'].value
        if isinstance(end_date, datetime.datetime):
            end_date = end_date.date()
        if run_description != thisRun_df['run_description'][0]:
            logging.critical('"run_description" differs between "runs" worksheet[%s] and  "clinical_costings_runs" table[%s]', run_description, thisRun_df['run_description'][0])
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if start_date != thisRun_df['start_date'][0]:
            logging.critical('"start_date" differs between "runs" worksheet[%s] and "clinical_costings_runs" table[%s]', start_date, thisRun_df['start_date'][0])
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if end_date != thisRun_df['end_date'][0]:
            logging.critical('"end_date" differs between "runs" worksheet[%s] and "clinical_costings_runs" table[%s]', end_date, thisRun_df['end_date'][0])
            logging.shutdown()
            sys.exit(d.EX_CONFIG)

    # Check if we are replacing an old run
    if not newRun:
        with d.Session() as session:
            session.execute(delete(d.metadata.tables['general_ledger_costs']).where(text(where)))
            session.execute(delete(d.metadata.tables['itemized_costs']).where(text(where)))
            session.execute(delete(d.metadata.tables['general_ledger_run_adjustments']).where(text(where)))
            session.execute(delete(d.metadata.tables['gl_attributes_run_adjustments']).where(text(where)))
            session.commit()
            logging.debug('general_ledger_costs, itemized_costs and event charges cleared')

    # Prepend the hospital_code and run code to the general ledger extract
    general_ledger_table_df.insert(0,'hospital_code', d.hospital_code)
    general_ledger_table_df.insert(1,'run_code', d.run_code)

    # Append the data to the general_ledger_costs table
    f.addTableData(general_ledger_table_df, 'general_ledger_costs')

    # Add the itemized costs
    for row in itemized_costs_df.itertuples():
        sheet = row.worksheet
        feeder_code = row.feeder_code
        table_df = sheet_table_df[sheet]

        # Prepend the hospital_code, run code and feeder_code
        table_df.insert(0,'hospital_code', d.hospital_code)
        table_df.insert(1,'run_code', d.run_code)
        table_df.insert(2,'feeder_code', feeder_code)

        # Append the data to the itemized_costs table
        f.addTableData(table_df, 'itemized_costs')

    # Add any general ledger run adjustments
    # Prepend the hospital_code and run code
    costAdjustments_df.insert(0,'hospital_code', d.hospital_code)
    costAdjustments_df.insert(1,'run_code', d.run_code)

    # Append the data to the general_ledger_run_adjustments table
    f.addTableData(costAdjustments_df, 'general_ledger_run_adjustments')

    # Add any general ledger attribute run adjustments
    # Prepend the hospital_code and run code
    attributeAdjustments_df.insert(0,'hospital_code', d.hospital_code)
    attributeAdjustments_df.insert(1,'run_code', d.run_code)

    # Append the data to the general_ledger_run_adjustments table
    f.addTableData(attributeAdjustments_df, 'gl_attributes_run_adjustments')

    logging.shutdown()
    sys.exit(d.EX_OK)
