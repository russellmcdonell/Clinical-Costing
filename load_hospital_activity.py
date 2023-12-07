
# pylint: disable=line-too-long, broad-exception-caught
'''
The load_hospital_activity script loads the patient activity data for a hospital
into the Clinical Costing tables, from an Excel workbook.

    SYNOPSIS:  
    $ python load_hospital_activity.py  
        [-I inputDir|--inputDir=inputDir] [-i inputWorkbook|--inputWorkbook=inputWorkbook]  
        [-C configDir|--configDir=configDir] [-c configFile|--configFile=configFile]  
        [-D DatabaseType|--DatabaseType=DatabaseType]  
        [-s server|--server=server]  
        [-u username|--username=username] [-p password|--password=password]  
        [-d databaseName|--databaseName=databaseName]  
        [-v loggingLevel|--verbose=logingLevel]  
        [-L logDir|--logDir=logDir] [-l logfile|--logfile=logfile]
<br/>

    OPTIONS  
    -I inputDir|--inputDir=inputDir  
    The directory containing the Excel workbook which contains the hospital patient activity data to be loaded.  

    -i inputWorkbook|--inputWorkbook=inputWorkbook  
    The Excel workbook which contains the hospital patient activity data to be loaded.  

    -C configDir|--configDir=configDir  
    The directory containing the database connection configuration file (default='databaseConfig')

    -c configFile|--configFile=configFile  
    The database connection configuration file (default=clinical_costing.json) which has the default database values for each Database Type.
    These can be overwritten using command line options.

    -D DatabaseType|--DatabaseType=DatabaseType  
    The type of database [choice:MSSQL/MySQL]

    -s server|--server=server]  
    The address of the database server

    -u userName|--userName=userName]  
    The user name require to access the database

    -p password|--userName=userName]  
    The user password require to access the database

    -d databaseName|--databaseName=databaseName]  
    The name of the database

    -v loggingLevel|--verbose=loggingLevel  
    Set the level of logging that you want.

    -O logDir|--logDir=logDir  
    The directory where the log file will be created.

    -o logfile|--logfile=logfile  
    The name of a log file where you want all messages captured.


THE MAIN CODE  
Start by parsing the command line arguements and setting up logging.  
Connect to the database and read the configuration from an Excel workbook.  
Then check that  configuration and load the data.
'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, unspecified-encoding

import os
import sys
import argparse
import logging
import datetime
import pandas as pd
from sqlalchemy import text, delete
from openpyxl import load_workbook
import functions as f
import data as d


# Define the required worksheets and associated database tables
requiredSheets = {
    'Inpat episode details': 'inpat_episode_details',
    'Inpat admissions': 'inpat_admissions',
    'Inpat discharges': 'inpat_discharges',
    'Inpat patient location': 'inpat_patient_location',
    'Inpat theatre details': 'inpat_theatre_details',
    'Clinic activity details': 'clinic_activity_details',
    'ED episode details': 'ed_episode_details',
    'ED admissions': 'ed_admissions',
    'ED discharges': 'ed_discharges'
}


if __name__ == '__main__':
    '''
    The main code
    Start by parsing the command line arguements and setting up logging.
    Then check the Excel workbook - it should have one sheet for the hospital code/name
    and one sheet for each hospital configuration table, in the Clinical Costing database.
    And each sheet should have a header row with headings that match the names of the columns in the associated database table.
    '''

    # Save the program name
    progName = sys.argv[0]
    progName = progName[0:-3]        # Strip off the .py ending

    # set the options
    parser = argparse.ArgumentParser(description='Load a Clinical Costing Model')
    parser.add_argument('-I', '--inputDir', dest='inputDir', default='.',
                        help='The directory containing the Excel workbook which contains the hospital patient activity data to be loaded.')
    parser.add_argument('-i', '--inputWorkbook', dest='inputWorkbook',
                        default='.', help='The name of the Excel workbook containing the hospital patient activity data to be loaded')
    f.addCommonArguments(parser)      # Add the common command line arguments
    args = parser.parse_args()

    # Parse the command line options
    args = parser.parse_args()
    inputDir = args.inputDir
    inputWorkbook = args.inputWorkbook
    configDir = args.configDir
    configFile = args.configFile
    DatabaseType = args.DatabaseType
    server = args.server
    username = args.username
    password = args.password
    databaseName = args.databaseName
    logDir = args.logDir
    logFile = args.logFile
    loggingLevel = args.verbose

    # Set up logging
    f.setupLogging(progName, logDir, logFile, loggingLevel)

    # Read in the configuration file - which must exist if required - and create the database engine
    f.createEngine(configDir, configFile, DatabaseType, server, username, password, databaseName)

    # Load the workbook
    wb = load_workbook(os.path.join(inputDir, inputWorkbook))

    # Check the 'hospital' worksheet
    table_df = f.checkWorksheet(wb, 'hospital', 'hospitals', [])

    # Get the hospital_code from the 'hospital' worksheet
    ws = wb['hospital']
    d.hospital_code = ws['A2'].value        # First (and only) hospital_code in the list

    # Check if this is a new hospital code, or upgraded configuration of an existing hospital
    hospitals_df = pd.read_sql_query(text('SELECT hospital_code FROM hospitals'), d.engine.connect())
    hospitals = hospitals_df.values.tolist()      # convert rows/columns to a list of lists (will be [[hospital_code]] )
    if not [d.hospital_code] in hospitals:
        logging.critical('hospital (%s) no in table "hospitals"', d.hospital_code)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)

    # Check the 'run' worksheet
    table_df = f.checkWorksheet(wb, 'run', 'clinical_costing_runs', ['hospital_code'])

    # Get the run_code from the 'run' worksheet
    ws = wb['run']
    d.run_code = ws['A2'].value        # First (and only) run_code in the list

    # Check if this is a run_code exists
    runs_df = pd.read_sql_query(text('SELECT run_code FROM clinical_costing_runs'), d.engine.connect())
    runs = runs_df.values.tolist()      # convert rows/columns to a list of lists (will be [[run_code]] )
    newRun = not [d.run_code] in runs

    # If this is a new run_code then add it to the table
    table_df = table_df.truncate(after=0)       # We only want the first row
    if newRun:
        # Prepend the hospital code and create a new run record
        table_df.insert(0,'hospital_code', d.hospital_code)
        table_df.to_sql('clinical_costing_runs', d.engine, if_exists='append', index=False)
    else:       # Check that this is the same run
        runs_df = pd.read_sql_query(text('SELECT * FROM clinical_costing_runs'), d.engine.connect())
        thisRun_df = runs_df[runs_df['run_code'] == d.run_code]
        run_description = ws['B2'].value
        start_date = ws['C2'].value
        if isinstance(start_date, datetime.datetime):
            start_date = start_date.date()
        end_date = ws['D2'].value
        if isinstance(end_date, datetime.datetime):
            end_date = end_date.date()
        if run_description != thisRun_df['run_description'][0]:
            logging.critical('run_description differs in "clinical_costings_runs" table (%s) and "runs" worksheet (%s)', thisRun_df['run_description'][0], run_description)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if start_date != thisRun_df['start_date'][0]:
            logging.critical('start_date differs in "clinical_costings_runs" table (%s) and "runs" worksheet (%s)', thisRun_df['start_date'][0], start_date)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if end_date != thisRun_df['end_date'][0]:
            logging.critical('end_date differs in "clinical_costings_runs" table (%s) and "runs" worksheet (%s)', thisRun_df['end_date'][0], end_date)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)

    # Create the general "where" clause
    where = 'hospital_code = "' + d.hospital_code + '" AND run_code = "' + d.run_code + '"'

    # Check that the required configuration data worksheets
    sheet_table_df = {}
    for sheet, table in requiredSheets.items():
        sheet_table_df[sheet] = f.checkWorksheet(wb, sheet, table, ['hospital_code', 'run_code'])

    # Check if we are replacing an old run
    if not newRun:
        with d.Session() as session:
            for sheet, table in reversed(requiredSheets.items()):
                session.execute(delete(d.metadata.tables[table]).where(text(where)))
            session.commit()

    # Add the patient activity data to the database tables
    for sheet, table in requiredSheets.items():
        table_df = sheet_table_df[sheet]

        # Prepend the hospital_code and run code
        table_df.insert(0,'hospital_code', d.hospital_code)
        table_df.insert(1,'run_code', d.run_code)

        # Append the data to the itemized_costs table
        f.addTableData(table_df, table)
