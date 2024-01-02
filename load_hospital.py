
# pylint: disable=line-too-long, broad-exception-caught
'''
The load_hospital script loads the configuration data for a hospital
into the Clinical Costing tables, from an Excel workbook.

    SYNOPSIS
    $ python load_hospital.py
        [-D DatabaseType|--DatabaseType=DatabaseType]
        [-C configDir|--configDir=configDir] [-c configFile|--configFile=configFile]
        [-I inputDir|--inputDir=inputDir] [-i inputWorkbook|--inputWorkbook=inputWorkbook]
        [-s server|--server=server]
        [-u username|--username=username] [-p password|--password=password]
        [-d databaseName|--databaseName=databaseName]
        [-v loggingLevel|--verbose=logingLevel]
        [-L logDir|--logDir=logDir] [-l logfile|--logfile=logfile]
<br/>

    REQUIRED
    -D databaseType|--databaseType=databaseType
    The type of database [eg:MSSQL/MySQL]


    OPTIONS
    -C configDir|--configDir=configDir
    The directory where the database connection configuration file can be found (default='databaseConfig')

    -c configFile|--configFile=configFile
    The database connection configuration file (default=clinical_costing.json) which has the default database values for each Database Type.
    These can be overwritten using command line options.

    -I inputDir|--inputDir=inputDir
    The directory containing the Excel workbook containing the hospital configuration data to be loaded.

    -i inputWorkbook|--inputWorkbook=inputWorkbook
    The Excel workbook which contains the hospital configuration data to be loaded.

    -s server|--server=server]
    The address of the database server

    -u userName|--userName=userName]
    The user name require to access the database

    -p password|--userName=userName]
    The user password require to access the database

    -d databaseName|--databaseName=databaseName]
    The name of the database

    -v loggingLevel|--verbose=loggingLevel
    Set the level of logging that you want.

    -O logDir|--logDir=logDir
    The directory where the log file will be created (default=".").

    -o logfile|--logfile=logfile
    The name of a log file where you want all messages captured.


THE MAIN CODE
Start by parsing the command line arguements and setting up logging.
Connect to the database and read the configuration from an Excel workbook.
Then check that configuration and load the data.
'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, unspecified-encoding

import os
import sys
import argparse
import logging
import pandas as pd
from sqlalchemy import text
from openpyxl import load_workbook
import functions as f
import data as d


# Define the required worksheets and associated database tables
requiredSheets = [
    {
        'feeder types': 'feeder_types'
    }, {
        'departments': 'departments',
        'cost types': 'cost_types',
        'services': 'services',
        'wards': 'wards',
        'theatres': 'theatres',
        'clinics': 'clinics',
        'clinicians': 'clinicians',
        'feeders': 'feeders'
    }
]


if __name__ == '__main__':
    '''
    The main code
    Start by parsing the command line arguements and setting up logging.
    Then check the Excel workbook - it should have one sheet for the hospital code/name
    and one sheet for each hospital configuration table, in the Clinical Costing database.
    And each sheet should have a header row with headings that match the names of the columns in the associated database table.
    '''

    # Save the program name
    progName = sys.argv[0]
    progName = progName[0:-3]        # Strip off the .py ending

    # Set the options
    parser = argparse.ArgumentParser(description='Load a Clinical Costing Model')
    parser.add_argument('-I', '--inputDir', dest='inputDir', default='.',
                        help='The directory containing the Excel workbook which contains the hospital patient activity data to be loaded.')
    parser.add_argument('-i', '--inputWorkbook', dest='inputWorkbook',
                        default='.', help='The name of the Excel workbook containing the hospital patient activity data to be loaded')
    f.addCommonArguments(parser)      # Add the common command line arguments
    args = parser.parse_args()

    # Parse the command line options
    args = parser.parse_args()
    inputDir = args.inputDir
    inputWorkbook = args.inputWorkbook
    configDir = args.configDir
    configFile = args.configFile
    DatabaseType = args.DatabaseType
    server = args.server
    username = args.username
    password = args.password
    databaseName = args.databaseName
    logDir = args.logDir
    logFile = args.logFile
    loggingLevel = args.verbose

    # Set up logging
    f.setupLogging(progName, logDir, logFile, loggingLevel)
    logging.debug('Logging set up')

    # Read in the configuration file - which must exist if required - and create the database engine
    f.createEngine(configDir, configFile, DatabaseType, server, username, password, databaseName)

    # Load the workbook
    wb = load_workbook(os.path.join(inputDir, inputWorkbook))

    # Check the 'hospital' worksheet
    table_df = f.checkWorksheet(wb, 'hospital', 'hospitals', [])

    # Get the hospital_code from the worksheet
    ws = wb['hospital']
    d.hospital_code = ws['A2'].value        # First (and only) hospital_code in the list

    # Check if this is a new hospital code, or upgraded configuration of an existing hospital
    hospitals_df = pd.read_sql_query(text('SELECT hospital_code FROM hospitals'), d.engine.connect())
    hospitals = hospitals_df.values.tolist()      # convert rows/columns to a list of lists (will be [[hospital_code]] )
    newHospital = not [d.hospital_code] in hospitals

    # If this is a new hospital, then add it to the table
    if newHospital:
        table_df = table_df.truncate(after=0)       # We only want the first row
        table_df.to_sql('hospitals', d.engine, if_exists='append', index=False)

    # Add this configuation data to the database
    for theseSheets in requiredSheets:
        for sheet, table in theseSheets.items():
            # Check the worksheet
            table_df = f.checkWorksheet(wb, sheet, table, ['hospital_code'])

            # Prepend the hospital code
            table_df.insert(0,'hospital_code', d.hospital_code)

            # Add the data to the database
            f.addTableData(table_df, table)

    logging.shutdown()
    sys.exit(d.EX_OK)
