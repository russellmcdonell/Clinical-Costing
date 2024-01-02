
# pylint: disable=line-too-long, broad-exception-caught
'''
The load_model script loads the clinical costing model configuration data, for a specific costing model
into the Clinical Costing tables, from an Excel workbook

    SYNOPSIS
    $ python load_model.py
        [-I inputDir|--inputDir=inputDir] [-i inputWorkbook|--inputWorkbook=inputWorkbook]
        [-C configDir|--configDir=configDir] [-c configFile|--configFile=configFile]
        [-D DatabaseType|--DatabaseType=DatabaseType]
        [-s server|--server=server]
        [-u username|--username=username] [-p password|--password=password]
        [-d databaseName|--databaseName=databaseName]
        [-v loggingLevel|--verbose=logingLevel]
        [-L logDir|--logDir=logDir] [-l logfile|--logfile=logfile]
<br/>

    OPTIONS
    -I inputDir|--inputDir=inputDir
    The directory containing the Excel workbook containing the model to be loaded.

    -i inputWorkbook|--inputWorkbook=inputWorkbook
    The Excel workbook containing the model configuration data, for this hospital, to be loaded.

    -C configDir|--configDir=configDir
    The directory where the database connection configuration file can be found (default='databaseConfig')

    -c configFile|--configFile=configFile
    The database connection configuration file (default=clinical_costing.json) which has the default database values for each Database Type.
    These can be overwritten using command line options.

    -D DatabaseType|--DatabaseType=DatabaseType
    The type of database [choice:MSSQL/MySQL]

    -s server|--server=server]
    The address of the database server

    -u userName|--userName=userName]
    The user name require to access the database

    -p password|--userName=userName]
    The user password require to access the database

    -d databaseName|--databaseName=databaseName]
    The name of the database

    -v loggingLevel|--verbose=loggingLevel
    Set the level of logging that you want.

    -O logDir|--logDir=logDir
    The directory where the log file will be created (default=".").

    -o logfile|--logfile=logfile
    The name of a log file where you want all messages captured.


THE MAIN CODE
Start by parsing the command line arguements and setting up logging.
Connect to the database and read the configuration from an Excel workbook.
Then check that configuration and load the data.
'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, unspecified-encoding

import os
import sys
import argparse
import logging
import pandas as pd
from sqlalchemy import text
from openpyxl import load_workbook
import functions as f
import data as d


# Define the required worksheets and associated database tables
requiredSheets = [
    {
        'event type codes': 'event_type_codes',
        'event source codes': 'event_source_codes',
    }, {
         'mapping types': 'mapping_types',
         'feeder model': 'feeder_model',
        'event class codes': 'event_class_codes',
        'event attribute codes': 'event_attribute_codes',
        'event subroutines': 'event_subroutines',
        'ward attributes': 'ward_attributes',
        'clinic attributes': 'clinic_attributes',
        'general ledger attribute codes': 'general_ledger_attribute_codes',
        'distribution codes': 'distribution_codes'
    }, {
       'general ledger mapping': 'general_ledger_mapping',
        'department grouping': 'department_grouping',
        'cost type grouping': 'cost_type_grouping',
        'department cost type grouping': 'department_cost_type_grouping',
        'event codes': 'event_codes',
        'event attributes': 'event_attributes',
        'general ledger attributes': 'general_ledger_attributes',
        'general ledger disbursement': 'general_ledger_disbursement',
        'general ledger distribution': 'general_ledger_distribution'
    }
]

if __name__ == '__main__':
    '''
    The main code
    Start by parsing the command line arguements and setting up logging.
    Then check the Excel workbook - it should have one sheet for the hospital code/name, one sheet for the model code/name
    and one sheet for each model configuration table, in the Clinical Costing database.
    And each sheet should have a header row with headings that match the names of the columns in the associated database table.
    '''

    # Save the program name
    progName = sys.argv[0]
    progName = progName[0:-3]        # Strip off the .py ending

    # Set the options
    parser = argparse.ArgumentParser(description='Load a Clinical Costing Model')
    parser.add_argument('-I', '--inputDir', dest='inputDir', default='.',
                        help='The directory containing the Excel workbook which contains the hospital patient activity data to be loaded.')
    parser.add_argument('-i', '--inputWorkbook', dest='inputWorkbook',
                        default='.', help='The name of the Excel workbook containing the hospital patient activity data to be loaded')
    f.addCommonArguments(parser)      # Add the common command line arguments
    args = parser.parse_args()

    # Parse the command line options
    args = parser.parse_args()
    inputDir = args.inputDir
    inputWorkbook = args.inputWorkbook
    configDir = args.configDir
    configFile = args.configFile
    DatabaseType = args.DatabaseType
    server = args.server
    username = args.username
    password = args.password
    databaseName = args.databaseName
    logDir = args.logDir
    logFile = args.logFile
    loggingLevel = args.verbose

    # Set up logging
    f.setupLogging(progName, logDir, logFile, loggingLevel)
    logging.debug('Logging set up')

    # Read in the configuration file - which must exist if required - and create the database engine
    f.createEngine(configDir, configFile, DatabaseType, server, username, password, databaseName)

    # Load the workbook
    wb = load_workbook(os.path.join(inputDir, inputWorkbook))

    # Check the 'hospital' worksheet
    table_df = f.checkWorksheet(wb, 'hospital', 'hospitals', [])

    # Get the hospital_code from the worksheet
    ws = wb['hospital']
    d.hospital_code = ws['A2'].value        # First (and only) hospital_code in the list

    # Check if this is hospital code exists in the database
    hospitals_df = pd.read_sql_query(text('SELECT hospital_code FROM hospitals'), d.engine.connect())
    hospitals = hospitals_df.values.tolist()      # convert rows/columns to a list of lists (will be [[hospital_code]] )
    haveHospital = [d.hospital_code] in hospitals
    if not haveHospital:
        logging.critical('No hospital_code %s exists in the Clinical Costing database')
        logging.critical('Must load the hospital departments and cost types before you can model the hospitals costs')
        logging.shutdown()
        sys.exit(d.EX_DATAERR)

    # Check that the 'model' worksheet exits
    table_df = f.checkWorksheet(wb, 'model', 'models', ['hospital_code'])

    # Get the model_code from the worksheet
    ws = wb['model']
    d.model_code = ws['A2'].value        # First (and only) model_code in the list

    # Check if this is a new model code, or upgraded configuration of an existing model
    models_df = pd.read_sql_query(text(f'SELECT model_code FROM models WHERE hospital_code = "{d.hospital_code}"'), d.engine.connect())
    models = models_df.values.tolist()      # convert rows/columns to a list of lists (will be [[model_code]] )
    newModel = not [d.model_code] in models

    # If this is a new model, then add it to the table
    if newModel:
        table_df = table_df.truncate(after=0)       # We only want the first row
        table_df.insert(0,'hospital_code', d.hospital_code)
        table_df.to_sql('models', d.engine, if_exists='append', index=False)

    # Add the first set of configuation data to the database
    theseSheets = requiredSheets[0]
    for sheet, table in theseSheets.items():
        # Check the worksheet
        table_df = f.checkWorksheet(wb, sheet, table, ['hospital_code', 'model_code'])

        # Prepend the hospital code
        table_df.insert(0,'hospital_code', d.hospital_code)
        table_df.insert(1,'model_code', d.model_code)

        # Add the data to the database
        f.addTableData(table_df, table)

    # Now use the hospital's feeder configuration data
    # to add codes to event_class_codes, event_attribute_code, distribution_codes and event_codes
    feeders_df = pd.read_sql_query(text('SELECT * FROM feeders WHERE hospital_code = "' + d.hospital_code + '"'), d.engine.connect())
    event_class_codes_df = feeders_df[['hospital_code', 'event_class_code', 'event_class_seq', 'feeder_description']]
    event_class_codes_df = event_class_codes_df.rename(columns={'feeder_description': 'event_class_description'})
    event_class_codes_df.insert(1, 'model_code', d.model_code)
    f.addTableData(event_class_codes_df, 'event_class_codes')
    event_codes_df = feeders_df[['hospital_code', 'feeder_code', 'feeder_description']]
    event_codes_df = event_codes_df.rename(columns={'feeder_code': 'event_attribute_code', 'feeder_description': 'event_attribute_description'})
    event_codes_df.insert(1, 'model_code', d.model_code)
    f.addTableData(event_codes_df, 'event_attribute_codes')
    event_codes_df = event_codes_df.rename(columns={'event_attribute_code': 'distribution_code', 'event_attribute_description': 'distribution_description'})
    f.addTableData(event_codes_df, 'distribution_codes')
    event_codes_df = feeders_df[['hospital_code', 'feeder_code', 'event_class_code', 'feeder_description']]
    event_codes_df = event_codes_df.rename(columns={'feeder_code': 'event_code', 'feeder_description': 'event_description'})
    event_codes_df.insert(1, 'model_code', d.model_code)
    event_codes_df.insert(2,'event_type_code', 'other')
    event_codes_df.insert(4,'event_source_code', 'Invoice')
    f.addTableData(event_codes_df, 'event_codes')

    # Add the remaining configuation data to the database
    for theseSheets in requiredSheets[1:]:
        for sheet, table in theseSheets.items():
            # Check the worksheet
            table_df = f.checkWorksheet(wb, sheet, table, ['hospital_code', 'model_code'])

            # Prepend the hospital code
            table_df.insert(0,'hospital_code', d.hospital_code)
            table_df.insert(1,'model_code', d.model_code)

            # Special case event_attributes, where event_where can be blank, but not null
            if table == 'event_attributes':
                table_df['event_what'].fillna('', inplace=True)

            # Add the data to the database
            f.addTableData(table_df, table)

    logging.shutdown()
    sys.exit(d.EX_OK)
